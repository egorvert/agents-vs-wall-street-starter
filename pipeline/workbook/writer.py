"""Workbook writer: final.json -> submission/<output_file>.xlsx.

Loads the organiser template, writes forecast values into column C next to each
exact metric label, touches nothing else. Labels/units come from
schemas/metrics.json and must match the template verbatim — a mismatch raises.
"""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent.parent
TEMPLATES = REPO / "challenge" / "templates"
SUBMISSION = REPO / "submission"


def load_manifest() -> dict:
    with open(REPO / "schemas" / "metrics.json") as f:
        return json.load(f)["companies"]


def write_workbook(company_id: str, finals: list[dict]) -> Path:
    info = load_manifest()[company_id]
    by_id = {f["metric_id"]: f for f in finals}
    wb = openpyxl.load_workbook(TEMPLATES / info["output_file"])
    ws = wb["Summary"]

    label_rows = {}
    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if isinstance(cell.value, str):
            label_rows[cell.value] = cell.row

    for m in info["metrics"]:
        label, mid = m["label"], m["metric_id"]
        if label not in label_rows:
            raise ValueError(f"{company_id}: label {label!r} not found in template Summary sheet")
        if mid not in by_id:
            raise ValueError(f"{company_id}: no final value for {mid} — never leave a metric blank")
        r = label_rows[label]
        unit_in_sheet = ws.cell(row=r, column=2).value
        if unit_in_sheet != m["unit"]:
            raise ValueError(f"{company_id} {mid}: template unit {unit_in_sheet!r} != manifest {m['unit']!r}")
        ws.cell(row=r, column=3).value = float(by_id[mid]["value"])

    SUBMISSION.mkdir(exist_ok=True)
    out = SUBMISSION / info["output_file"]
    wb.save(out)
    return out
